from compartidos import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
import os


# Especifica el directorio donde se guardara el archivo Excel con las posibles soluciones
# Estas se encuentran ordenas de mejor a peor
directorio_resultados = "resultados"

# Verifica si el directorio existe, de no existir se creara
if not os.path.exists(directorio_resultados):
    os.makedirs(directorio_resultados)


paso = input("¿Que punto desea realizar?\n1. Exhaustivo\n2. Heuristica\n3. Exhaustivo\n4. Heuristica\n\n")
while (paso != "1" and paso != "2" and paso != "3" and paso != "4"):
    paso = input("¿Que punto desea realizar?\n1. Exhaustivo\n2. Heuristica\n3. Exhaustivo\n4. Heuristica\n\n")

if (paso == "1" or paso == "2"):
    objetos: list = [
        Objeto(150, 0, 20),
        Objeto(325, 0, 40),
        Objeto(600, 0, 50),
        Objeto(805, 0, 36),
        Objeto(430, 0, 25),
        Objeto(1200, 0, 64),
        Objeto(770, 0, 54),
        Objeto(60, 0, 18),
        Objeto(930, 0, 46),
        Objeto(353, 0, 28),
    ]
    # Capacidad o volumen maximo soportado por la mochila, dato dado por el enunciado
    capacidad = 4200

if paso == "3" or paso == "4":
    # Datos de los objetos dados por el enunciado
    objetos: list = [
        Objeto(0, 1800, 72),
        Objeto(0, 600, 36),
        Objeto(0, 1200, 60)
    ]
    # Capacidad maxima soportada por la mochila, dato dado por el enunciado
    capacidad = 3000

# Dado que luego el orden de los objetos se altera, 
# necesito una copia para poder operar sobre el orden original mas adelante
copiaObjetos = objetos.copy()


if paso == "1" or paso == "3":
    # Lista de la solucion, cada posicion de esta lista coincide con la posicion del objeto en la lista objetos
    # En una etapa inicial o intermedia todos o algunos de los elementos de la lista seran None (incertidumbre)
    # Si meto un objeto a la mochila se simboliza con un True si no lo guardo se simboliza con un False
    # Por ejemplo si mi lista solucion es [False, False, True, True, False, False, False, False, True, True] 
    # significa que guarde en mi mochila los objetos objetos[2], objetos[3], objetos[8] y objetos[9]
    solucion = [None for i in range(len(objetos))]


    # En esta lista se guardaran todas las soluciones finales validas al problema
    # Esta lista es sometida a un ordenamiento para ordenar de la mejor a la peor
    # solucion segun el valor total de los objetos guardados en la mochila
    #Si no se aplican restricciones habria 2^n soluciones, siendo "n" la cantidad de objetos disponibles
    #El numero decrece mucho si se aplican restricciones
    soluciones = []


    buscarSolucion(solucion, objetos, capacidad, soluciones)
    ordenarSoluciones(soluciones, objetos)

else:
    soluciones = buscarSolucionHeu(objetos, capacidad)


# Guardado de archivo xls

# Crear un nuevo libro de Excel
libro_excel = Workbook()
hoja_excel = libro_excel.active
hoja_excel.title = f"Soluciones de la Mochila"

bordeDelgado = Side(border_style="thin", color="000000")

# Instrucciones generales para todos los pasos
# Encabezados
hoja_excel['A1'] = f"Paso {paso}"
hoja_excel['A1'].font = Font(bold=True)
alinearCelda(hoja_excel['A1'])
ponerBorde(hoja_excel['A1'], bordeDelgado)

hoja_excel['A2'].font = Font(bold=True)
alinearCelda(hoja_excel['A2'])
ponerBorde(hoja_excel['A2'], bordeDelgado)

hoja_excel['B2'].font = Font(bold=True)
alinearCelda(hoja_excel['B2'])
ponerBorde(hoja_excel['B2'], bordeDelgado)

#Exclusivo paso 1 y 3
if paso == "1" or paso == "3":
    # Tamaño columnas
    hoja_excel.column_dimensions["A"].width = 55
    hoja_excel.column_dimensions["B"].width = 15

    # Encabezados
    hoja_excel['A2'] = "Solución"
    hoja_excel['B2'] = "Valor total"

    hoja_excel.merge_cells('A1:B1')

    # Llenar la hoja de cálculo con las soluciones y sus valores
    for indice, solucion in enumerate(soluciones):
        fila = indice + 3  # Empezar desde la fila 3, ya que la primera fila contiene encabezados
        
        #En esta variable guardo objetos almacenados
        objetosMochila = []

        for i in range(len(solucion["solucion"])):
            if solucion["solucion"][i]:
                objetosMochila.append(str(i + 1))

        # Segun haya un objeto almacenado, muchos o ninguno
        # muestro en el excel el mensaje correspondiente
        if len(objetosMochila) == 0:
            hoja_excel[f'A{fila}'] = "No se almacenan objetos"
        elif len(objetosMochila) == 1:
            hoja_excel[f'A{fila}'] = "Se almacena el objeto: " + "".join(objetosMochila)
        else:
            hoja_excel[f'A{fila}'] = "Se almacenan los objetos: " + ", ".join(objetosMochila)

        hoja_excel[f'B{fila}'] = solucion["total"]

        ponerBorde(hoja_excel[f'A{fila}'], bordeDelgado)
        #alinearCelda(hoja_excel[f'A{fila}'])

        ponerBorde(hoja_excel[f'B{fila}'], bordeDelgado)
        alinearCelda(hoja_excel[f'B{fila}'])

#Exclusivo paso 2 y 4
else:
    # Tamaño columnas
    hoja_excel.column_dimensions["A"].width = 15
    hoja_excel.column_dimensions["B"].width = 15
    hoja_excel.column_dimensions["C"].width = 15

    # Encabezados
    # Segun lo que pida el ejercicio muestro los datos correspondientes
    if paso == "2":
        hoja_excel['A2'] = "Volumen"
    else:
        hoja_excel['A2'] = "Peso"

    hoja_excel['B2'] = "Valor"
    hoja_excel['C2'] = "Indice"
    hoja_excel['D2'] = "Objeto"

    hoja_excel.merge_cells('A1:D1')

    #Pongo formato a las celdas

    hoja_excel['C2'].font = Font(bold=True)
    alinearCelda(hoja_excel['C2'])
    ponerBorde(hoja_excel['C2'], bordeDelgado)

    hoja_excel['D2'].font = Font(bold=True)
    alinearCelda(hoja_excel['D2'])
    ponerBorde(hoja_excel['D2'], bordeDelgado)

    # Llenar la hoja de cálculo con las soluciones y sus valores
    for indice, solucion in enumerate(soluciones[0]):
        fila = indice + 3  # Empezar desde la fila 3, ya que la primera fila contiene encabezados

        #Segun lo que pide el ejercicio muestro el dato correspondiente
        if paso == "2":
            hoja_excel[f'A{fila}'] = solucion.volumen
        else:
            hoja_excel[f'A{fila}'] = solucion.peso

        hoja_excel[f'B{fila}'] = solucion.valor
        hoja_excel[f'C{fila}'] = solucion.indice

        #Pongo cual es el numero del objeto, para poderlo identificar
        hoja_excel[f'D{fila}'] = copiaObjetos.index(solucion) + 1

        ponerBorde(hoja_excel[f'A{fila}'], bordeDelgado)
        alinearCelda(hoja_excel[f'A{fila}'])

        ponerBorde(hoja_excel[f'B{fila}'], bordeDelgado)
        alinearCelda(hoja_excel[f'B{fila}'])

        ponerBorde(hoja_excel[f'C{fila}'], bordeDelgado)
        alinearCelda(hoja_excel[f'C{fila}'])

        ponerBorde(hoja_excel[f'D{fila}'], bordeDelgado)
        alinearCelda(hoja_excel[f'D{fila}'])

    # Añado el valor total de la solucion debajo de todo
    hoja_excel[f'A{(len(soluciones[0])+3)}'] = f"Valor Total: {soluciones[1]}"
    # Le pongo formato
    ponerBorde(hoja_excel[f'A{(len(soluciones[0])+3)}'], bordeDelgado)
    alinearCelda(hoja_excel[f'A{(len(soluciones[0])+3)}'])
    hoja_excel[f'A{(len(soluciones[0])+3)}'].font = Font(bold=True)

    hoja_excel.merge_cells(f'A{(len(soluciones[0])+3)}:D{(len(soluciones[0])+3)}')


# Guardar el libro de Excel en el directorio de resultados
nombre_archivo_excel = f"Soluciones Mochila Paso {paso}.xlsx"
ruta_archivo_excel = os.path.join(directorio_resultados, nombre_archivo_excel)
libro_excel.save(ruta_archivo_excel)

print(f"\nSe han guardado las soluciones en '{ruta_archivo_excel}'")